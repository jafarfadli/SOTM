import shutil
import os
import re
from openpyxl import load_workbook
from typing import List

# fungsi
def col_letter_to_number(col_letter: str) -> int:
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

def transpose_data(raw_data: List[List[str]]) -> List[List[str]]:
    return [list(row) for row in zip(*raw_data)]

def get_data_excel(source_path: str, sheet_name: str, range_cell: str) -> List[List[str]]:
    match = re.match(r'([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)', range_cell.upper())    
    start_col_letter, start_row, end_col_letter, end_row = match.groups()
    
    start_col = col_letter_to_number(start_col_letter)
    end_col = col_letter_to_number(end_col_letter)
    start_row = int(start_row)
    end_row = int(end_row)
    
    workbook = load_workbook(source_path, data_only=True)
    sheet = workbook[sheet_name]
    
    result = []
    for row in range(start_row, end_row + 1):
        row_data = []
        for col in range(start_col, end_col + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is None:
                row_data.append("")
            else:
                row_data.append(str(cell_value))
        result.append(row_data)
    workbook.close()
    return result

def edit_data_excel(source_path: str, sheet_name: str, start_cell: str, new_data: List[List[str]]):
    match = re.match(r'([A-Za-z]+)(\d+)', start_cell.upper())    
    start_col_letter, start_row = match.groups() 
    
    start_col = col_letter_to_number(start_col_letter)
    start_row = int(start_row)   

    workbook = load_workbook(source_path)
    
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    
    sheet = workbook[sheet_name]

    for i, row_data in enumerate(new_data):
        for j, cell_value in enumerate(row_data):
            row = start_row + i
            col = start_col + j
        
            if cell_value.isdigit():
                # int
                sheet.cell(row=row, column=col).value = int(cell_value)
            elif cell_value.replace('.', '', 1).isdigit() and cell_value.count('.') == 1:
                # float
                sheet.cell(row=row, column=col).value = float(cell_value)
            elif cell_value.lower() in ['true', 'false']:
                # bool
                sheet.cell(row=row, column=col).value = cell_value.lower() == 'true'
            else:
                # str
                sheet.cell(row=row, column=col).value = cell_value
    workbook.save(source_path)
    workbook.close()
    
def copy_file(source_path: str, destination_path: str) -> bool:
    shutil.copy2(source_path, destination_path)

def move_file(source_path: str, destination_path: str):
    shutil.move(source_path, destination_path)
    
# cek berkas
executable = True
semester = ''
tahun = ''
for filename in os.listdir('mata_kuliah'):
    identitas_semester = get_data_excel(f'mata_kuliah/{filename}', 'SUMMARY', 'B4:B5')
    if (semester == '' and tahun == ''):
        semester = identitas_semester[0][0].lower()
        tahun = identitas_semester[1][0]
    elif (semester.lower() != identitas_semester[0][0].lower() or tahun != identitas_semester[1][0]):
        print(f'ERROR: Terdapat kesalahan pengisian Semester/Tahun di {filename}')
        executable = False
        break
    
if executable:
    # transkrip semester
    cpl_semester = [['0' for i in range(10)] for j in range(4)]

    for filename in os.listdir('mata_kuliah'):
        cpl_mata_kuliah = get_data_excel(f'mata_kuliah/{filename}', 'SUMMARY', 'H7:Q10')
        for i in range(10):
            for j in range(4):
                cpl_semester[j][i] = str(int(cpl_semester[j][i]) + int(cpl_mata_kuliah[j][i]))
                
    copy_file('template/template_semester.xlsx','transkrip_semester.xlsx')
    edit_data_excel('transkrip_semester.xlsx', 'SUMMARY', 'B11', transpose_data(cpl_semester))
    edit_data_excel('transkrip_semester.xlsx', 'SUMMARY', 'B1', [[semester[0].upper()+semester[1:]],[tahun]])

    # transkrip mahasiswa
    all_data_mahasiswa = []
    for filename in os.listdir('mata_kuliah'):
        data_mahasiswa = get_data_excel(f'mata_kuliah/{filename}', 'SUMMARY', 'B11:Q90')
        data_mk = get_data_excel(f'mata_kuliah/{filename}', 'SUMMARY', 'B1:B2')
        for row in data_mahasiswa:
            if row[0] != '':
                all_data_mahasiswa += [row[0:2] + data_mk[0] + data_mk[1] + row[4:16]]

    unique_nim_nama = []
    for row in all_data_mahasiswa:
        if not (row[0] in unique_nim_nama):
            unique_nim_nama += [row[0:2]]

    for nim_nama in unique_nim_nama:
        specific_data_mahasiswa = []
        ip_semester = 0
        ip_dictionary = {'A':'4','AB':'3.5','B':'3','BC':'2.5','C':'2','D':'1','E':'0'}    
        for row in all_data_mahasiswa:
            if nim_nama[0] == row[0]:
                specific_data_mahasiswa += [row[2:]]
        for row in specific_data_mahasiswa:
            ip_semester += float(ip_dictionary[row[2]]) / len(specific_data_mahasiswa)
        ip_semester = round(ip_semester, 2)
        
        copy_file('template/template_mahasiswa.xlsx', f'mahasiswa/transkrip_{nim_nama[0]}.xlsx')
        edit_data_excel(f'mahasiswa/transkrip_{nim_nama[0]}.xlsx', 'SUMMARY', 'B11', specific_data_mahasiswa)
        edit_data_excel(f'mahasiswa/transkrip_{nim_nama[0]}.xlsx', 'SUMMARY', 'B1', [[nim_nama[1]],[nim_nama[0]],[semester[0].upper()+semester[1:]],[tahun],[str(ip_semester)]])
    print(f'BERHASIL')