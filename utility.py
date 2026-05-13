import shutil
import os
import re
from openpyxl import load_workbook
from typing import List
from copy import copy

# fungsi
def col_letter_to_number(col_letter: str) -> int:
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

def number_to_col_letter(col_number: int) -> str:
    result = ''
    while col_number > 0:
        col_number -= 1
        result = chr(col_number % 26 + ord('A')) + result
        col_number //= 26
    return result

def transpose_data(raw_data: List[List[str]]) -> List[List[str]]:
    return [list(row) for row in zip(*raw_data)]

def get_data_excel(source_path: str, sheet_name: str, range_cell: str) -> List[List[str]]:
    match = re.match(r'([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)', range_cell.upper())    
    start_col_letter, start_row, end_col_letter, end_row = match.groups()
    
    start_col = col_letter_to_number(start_col_letter)
    end_col = col_letter_to_number(end_col_letter)
    start_row = int(start_row)
    end_row = int(end_row)
    
    workbook = load_workbook(source_path, data_only=True)
    sheet = workbook[sheet_name]
    
    result = []
    for row in range(start_row, end_row + 1):
        row_data = []
        for col in range(start_col, end_col + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is None:
                row_data.append('')
            else:
                row_data.append(str(cell_value))
        result.append(row_data)
    workbook.close()
    return result

def edit_data_excel(source_path: str, sheet_name: str, start_cell: str, new_data: List[List[str]]):
    match = re.match(r'([A-Za-z]+)(\d+)', start_cell.upper())    
    start_col_letter, start_row = match.groups() 
    
    start_col = col_letter_to_number(start_col_letter)
    start_row = int(start_row)   

    workbook = load_workbook(source_path)
    
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    
    sheet = workbook[sheet_name]

    for i, row_data in enumerate(new_data):
        for j, cell_value in enumerate(row_data):
            row = start_row + i
            col = start_col + j
        
            if cell_value.isdigit():
                # int
                sheet.cell(row=row, column=col).value = int(cell_value)
            elif cell_value.replace('.', '', 1).isdigit() and cell_value.count('.') == 1:
                # float
                sheet.cell(row=row, column=col).value = float(cell_value)
            elif cell_value.lower() in ['true', 'false']:
                # bool
                sheet.cell(row=row, column=col).value = cell_value.lower() == 'true'
            else:
                # str
                sheet.cell(row=row, column=col).value = cell_value
    workbook.save(source_path)
    workbook.close()
    
def copy_data_excel(source_path: str, sheet_name: str, range_cell: str, target_cell: str):
    match = re.match(r'([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)', range_cell.upper())    
    start_col_letter, start_row, end_col_letter, end_row = match.groups()
    
    start_col = col_letter_to_number(start_col_letter)
    end_col = col_letter_to_number(end_col_letter)
    start_row = int(start_row)
    end_row = int(end_row)
    
    target_match = re.match(r'([A-Za-z]+)(\d+)', target_cell.upper())    
    target_col_letter, target_row = target_match.groups() 
    
    target_col = col_letter_to_number(target_col_letter)
    target_row = int(target_row)

    workbook = load_workbook(source_path)
    
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    
    sheet = workbook[sheet_name]
    
    for i, row in enumerate(range(start_row, end_row + 1)):
        for j, col in enumerate(range(start_col, end_col + 1)):
            source_cell = sheet.cell(row=row, column=col)
            
            target_row_pos = target_row + i
            target_col_pos = target_col + j
            
            target_cell_obj = sheet.cell(row=target_row_pos, column=target_col_pos)            
            target_cell_obj.value = source_cell.value
            
            if source_cell.has_style:
                # font
                if source_cell.font:
                    target_cell_obj.font = copy(source_cell.font)                
                # border
                if source_cell.border:
                    target_cell_obj.border = copy(source_cell.border)                
                # fill
                if source_cell.fill:
                    target_cell_obj.fill = copy(source_cell.fill)                
                # number format
                if source_cell.number_format:
                    target_cell_obj.number_format = source_cell.number_format                
                # alignment
                if source_cell.alignment:
                    target_cell_obj.alignment = copy(source_cell.alignment)
                
    workbook.save(source_path)
    workbook.close()

def copy_file(source_path: str, destination_path: str):
    shutil.copy2(source_path, destination_path)

def move_file(source_path: str, destination_path: str):
    shutil.move(source_path, destination_path)
    
def check_file_exists(file_path: str) -> bool:
    return os.path.isfile(file_path)